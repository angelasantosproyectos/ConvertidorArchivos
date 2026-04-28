import csv
import sys
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl no está instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    _TKINTER_AVAILABLE = True
except ImportError:
    _TKINTER_AVAILABLE = False


# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
VALID_MODALIDAD = {"200111", "200112", "200113"}
VALID_FORMA_PAGO = {"1", "2", "3"}
VALID_REGIMEN = {"F", "J"}

DATE_FORMATS = [
    "%d/%m/%Y",
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%m/%d/%Y",
]


# ---------------------------------------------------------------------------
# Diálogos de selección de archivo y opciones
# ---------------------------------------------------------------------------

def _init_tk():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    return root


def ask_excel_file():
    if not _TKINTER_AVAILABLE:
        return input("Introduce ruta del Excel: ").strip()

    root = _init_tk()
    path = filedialog.askopenfilename(
        title="Selecciona el archivo Excel",
        filetypes=[("Archivos Excel", "*.xlsx *.xlsm *.xls"), ("Todos", "*.*")],
    )
    root.destroy()

    if not path:
        print("No se seleccionó ningún archivo. Saliendo.")
        sys.exit(0)
    return path


def ask_delimiter():
    """Muestra un diálogo para elegir el delimitador: coma o punto y coma."""
    if not _TKINTER_AVAILABLE:
        while True:
            choice = input("Separador (1 = coma  ,   2 = punto y coma  ;): ").strip()
            if choice == "1":
                return ","
            elif choice == "2":
                return ";"
            print("Introduce 1 o 2.")

    result = {"delimiter": None}

    dialog = tk.Toplevel()
    dialog.title("Formato de separador")
    dialog.resizable(False, False)
    dialog.attributes("-topmost", True)

    # Centrar en pantalla
    dialog.update_idletasks()
    w, h = 360, 160
    x = (dialog.winfo_screenwidth() - w) // 2
    y = (dialog.winfo_screenheight() - h) // 2
    dialog.geometry(f"{w}x{h}+{x}+{y}")

    tk.Label(
        dialog,
        text="¿Con qué separador quieres generar el CSV?",
        font=("Arial", 11),
        wraplength=320,
        pady=12,
    ).pack()

    btn_frame = tk.Frame(dialog)
    btn_frame.pack(pady=8)

    def pick(delim):
        result["delimiter"] = delim
        dialog.destroy()

    tk.Button(
        btn_frame,
        text="  Coma  ( , )  —  estándar CSV",
        font=("Arial", 10),
        width=28,
        command=lambda: pick(","),
    ).grid(row=0, column=0, padx=6, pady=4)

    tk.Button(
        btn_frame,
        text="  Punto y coma  ( ; )  —  Excel",
        font=("Arial", 10),
        width=28,
        command=lambda: pick(";"),
    ).grid(row=1, column=0, padx=6, pady=4)

    dialog.protocol("WM_DELETE_WINDOW", lambda: sys.exit(0))
    dialog.grab_set()
    dialog.wait_window()

    return result["delimiter"]


def ask_csv_save_path():
    if not _TKINTER_AVAILABLE:
        path = input("Introduce ruta del CSV de salida: ").strip()
        if not path.lower().endswith(".csv"):
            path += ".csv"
        return path

    root = _init_tk()
    path = filedialog.asksaveasfilename(
        title="Guardar CSV como...",
        defaultextension=".csv",
        filetypes=[("Archivos CSV", "*.csv"), ("Todos", "*.*")],
    )
    root.destroy()

    if not path:
        print("No se seleccionó destino. Saliendo.")
        sys.exit(0)

    if not path.lower().endswith(".csv"):
        path += ".csv"
    return path


# ---------------------------------------------------------------------------
# Funciones de transformación
# ---------------------------------------------------------------------------

def extract_x_from_x_texto(value):
    if "-" in value:
        return value.split("-", 1)[0].strip()
    return value.strip()


def transform_regimen(value):
    part = extract_x_from_x_texto(value)
    return part if part in VALID_REGIMEN else "0"


def transform_modalidad(value):
    part = extract_x_from_x_texto(value)
    if part == "20013":
        part = "200113"
    return part if part in VALID_MODALIDAD else "0"


def transform_forma_pago(value):
    part = extract_x_from_x_texto(value)
    return part if part in VALID_FORMA_PAGO else "0"


def format_fecha(value):
    if value is None:
        return "0"
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%dT00:00:00.000+0100")
    raw = str(value).strip()
    if not raw:
        return "0"
    for fmt in DATE_FORMATS:
        try:
            dt = datetime.strptime(raw, fmt)
            return dt.strftime("%Y-%m-%dT00:00:00.000+0100")
        except ValueError:
            continue
    return "0"


def normalize_decimal_dot(value):
    """Para CSV con coma: decimales con punto  →  3,14  →  3.14"""
    return value.replace(",", ".")


def normalize_decimal_comma(value):
    """Para CSV con punto y coma: decimales con punto  →  3,14  →  3.14"""
    return value.replace(",", ".")


def cell_to_str(cell_value):
    if cell_value is None:
        return ""
    if hasattr(cell_value, "strftime"):
        return cell_value.strftime("%d/%m/%Y")
    return str(cell_value).strip()


def process_generic(value, use_semicolon):
    if not value:
        return "0"
    transformed = extract_x_from_x_texto(value)
    if use_semicolon:
        transformed = normalize_decimal_comma(transformed)
    else:
        transformed = normalize_decimal_dot(transformed)
    return transformed if transformed else "0"


# ---------------------------------------------------------------------------
# Lógica principal
# ---------------------------------------------------------------------------

def get_column_index(headers, name):
    name_lower = name.lower()
    for i, h in enumerate(headers):
        if h is not None and str(h).strip().lower() == name_lower:
            return i
    return None


def process_row(row, headers, col_map, use_semicolon):
    values = [cell_to_str(cell) for cell in row]
    result = []

    for i, raw in enumerate(values):
        col_name = headers[i].strip() if headers[i] else ""
        col_lower = col_name.lower()

        if col_lower == "regimen":
            result.append(transform_regimen(raw) if raw else "0")
        elif col_lower == "modalidad":
            result.append(transform_modalidad(raw) if raw else "0")
        elif col_lower == "forma_pago":
            result.append(transform_forma_pago(raw) if raw else "0")
        else:
            result.append(process_generic(raw, use_semicolon))

    fecha_idx = col_map.get("fec_efecto_spto")
    if fecha_idx is not None:
        result.append(format_fecha(row[fecha_idx]))
    else:
        result.append("0")

    return result


def find_data_start(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        non_empty = [c for c in row if c.value is not None]
        if len(non_empty) >= 3:
            return row[0].row
    return 1


def convert(excel_path, csv_path, delimiter):
    if not os.path.isfile(excel_path):
        print(f"Error: No se encontró el archivo '{excel_path}'.")
        sys.exit(1)

    use_semicolon = delimiter == ";"

    print(f"Leyendo '{excel_path}'...")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"Error al abrir el Excel: {e}")
        sys.exit(1)

    ws = wb.active
    header_row_num = find_data_start(ws)

    raw_headers = [cell.value for cell in ws[header_row_num]]
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]

    col_map = {}
    for key in ("regimen", "modalidad", "forma_pago", "fec_efecto_spto"):
        idx = get_column_index(headers, key)
        if idx is not None:
            col_map[key] = idx

    output_headers = headers + ["FEC_EFECTO_SPTO_FORMATEADA"]
    rows_written = 0

    # Con punto y coma añadimos la BOM para que Excel lo abra directamente
    encoding = "utf-8-sig" if use_semicolon else "utf-8"

    try:
        with open(csv_path, "w", newline="", encoding=encoding) as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerow(output_headers)

            for row in ws.iter_rows(min_row=header_row_num + 1, values_only=False):
                if all(cell.value is None for cell in row):
                    continue
                first_val = str(row[0].value or "")
                if first_val.startswith("["):
                    continue

                raw_values = tuple(cell.value for cell in row)
                if len(raw_values) < len(headers):
                    raw_values += (None,) * (len(headers) - len(raw_values))

                processed = process_row(raw_values, headers, col_map, use_semicolon)
                writer.writerow(processed)
                rows_written += 1

    except IOError as e:
        print(f"Error al escribir el CSV: {e}")
        sys.exit(1)

    delim_label = "punto y coma (;) — listo para Excel" if use_semicolon else "coma (,)"
    print(f"CSV generado correctamente: '{csv_path}' ({rows_written} filas, separador: {delim_label}).")

    if _TKINTER_AVAILABLE:
        root = _init_tk()
        messagebox.showinfo(
            "Conversión completada",
            f"CSV generado correctamente.\n\n"
            f"{rows_written} filas procesadas.\n"
            f"Separador: {delim_label}\n\n"
            f"Archivo: {csv_path}",
        )
        root.destroy()


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

def main():
    excel_path = ask_excel_file()
    print(f"Excel seleccionado: {excel_path}")

    delimiter = ask_delimiter()
    print(f"Separador elegido: '{delimiter}'")

    csv_path = ask_csv_save_path()
    print(f"CSV de salida: {csv_path}")

    convert(excel_path, csv_path, delimiter)


if __name__ == "__main__":
    main()